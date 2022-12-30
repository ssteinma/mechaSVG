import os, random, datetime, functools, math, sys, subprocess
try:
	import tkinter as tk
	from tkinter import ttk, filedialog, messagebox
except ImportError as message:
	print(message)
	print("Please make sure you have python3, tkinter and ttk installed")
	input("Press enter to leave")
	exit()

class Preferences:
	def __init__(self):
		################################ RELATIVELY STRAIGHTFORWARD MODIFICATIONS ######################################
		# number of structures for each path
		self.n_structures = 30
		self.n_connectors = 8
		self.n_comparers = 5
		#run command
		self.windons_command = "start inkscape.exe ./.E_profile.svg" # Please note .E_profile.svg is a hidden file!
		self.linux_command = "inkscape ./.E_profile.svg" # Please note .E_profile.svg is a hidden file!
		self.command_line = self.windons_command if os.name == "nt" else self.linux_command
		# SVG colors
		self.menu_a = ["steelblue","black","cornflowerblue","lightseagreen","darkblue","orangered","darkred","deeppink","darkgreen"]
		#SVG line widths
		self.menu_b = ["2","3","4","5","6"]
		#SVG line stiles
		self.svg_repl = {"full": "","dashed":'stroke-dasharray="10,10"',"dashed1":'stroke-dasharray="6,6"',"dashed2":'stroke-dasharray="4,4"',"dashed3":'stroke-dasharray="2,2"'}
		# Random PES generator
		self.trickster = True # Include random PES generator?
		self.name = "MechaSVG v 0.1.1"
		######################## YOU ARE PROBABLY BETTER OFF NOT MESSING WITH THE FOLLOWING ############################
		self.menu_c = list(self.svg_repl.keys())
		# TDI and TDTS placement corrections
		self.placement = {
			"Top"      :[[-37,-22,-7],[-22,-7,0]],
			 "Middle"  :[[-5,15,30],[-5,15,30]],
			 "Bottom"  :[[17,32,47],[17,32,0]]
			}
		self.alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZðƿþæœȝ"
		self.menu_d = list(self.placement.keys())
		self.menu_e = [self.alphabet[n] for n in range(8)] #Will change the number of Paths available
		self.menu_f = ["opt_{}".format(a.lower()) for a in self.menu_e]
		self.menu_g = ["tab_{}".format(a.lower()) for a in self.menu_e]
		self.menu_h = ["Path {}".format(a) for a in self.menu_e]
		self.menu_i = ["#{}".format(a) for a in self.menu_e]
		self.menu_z = [" ","TS","INT"]
		try:
			import openpyxl
			self.xlsx = True
		except ImportError:
			self.xlsx = False
		if self.xlsx:
			self.allowed_extensions = {
				"title": "Save state",
				"defaultextension": ".xlsx",
				"filetypes": [("Spreadsheet", ".xlsx"),  ("Text file", ".txt")]}
		else:
			self.allowed_extensions = {
				"title": "Save state",
				"defaultextension": ".txt",
				"filetypes": [("Text file", ".txt")]}
		filename = sys.argv[-1]
		ext = [".xlsx",".txt"] if self.xlsx else [".txt"]
		if os.path.isfile(filename) and any(filename.endswith(a) for a in ext):
			self.filename = filename

		i = []
		labels = ["XY",  "Y only",  "X,Y","Frame", "X only", "Borderless"]
		i.append("^            #^            #^            #┌───────────┐#             #             ")
		i.append("│_   _   _   #│_   _   _   #│_   _   _   #│_   _   _  │# _   _   _   # _   _   _   ")
		i.append("│ \_/ \_/ \_ #│ \_/ \_/ \_ #│ \_/ \_/ \_ #│ \_/ \_/ \_│#  \_/ \_/ \_ #  \_/ \_/ \_ ")
		i.append("└───────────>#˅            #˅  <───────> #└───────────┘# <─────────> #             ")
		self.image = {}
		for idx,label in zip(range(len(i[0].split("#"))),labels):
			"XY", "X,Y", "X only", "Y only", "Frame", "Borderless"
			self.image[label] = "\n".join([[a.split("#") for a in i][n][idx] for n in range(4)])
		self.hints = [
			"Hint: svg graphs can be easily eddited with freely available svg editors like Inkscape.",
			"Hint: mechaSVG can read and save .xlsx files for MS excel or libreoffice.",
			"Hint: mechaSVG can read and save .txt files.",
			"Hint: Are labels overlaid? Try changing the 'Alignment' settings or 'Vertical' displacements.",
			"Hint: On the 'Horizontal' tab you can edit the horizontal length of the graph.",
			"Hint: The 'TS mark' option on 'Labels' only modifies labels if the corresponding structure 'Type' is set to 'TS'.",
			"Hint: On the 'Main' tab you can chosse wheter to plot Free energy or Enthalpy via the option 'Use Enthalpy instead of free energy'.",
			"Hint: mechaSVG ignores the type of data that is not being plotted i.e., if you are plotting Free energy, the values for Henthalpy --need not-- be numeric.",
			"Hint: Inkscape allows one to save '.svg' files as '.emf' files. '.emf' files can be imported to some of the mainstream chemical structure drawing programs with minimal loss in image quality.",
			"Hint: When you click on the preview buttons 'Command' or 'Default', mechaSVG will save a hidden file on the current directory named '.E_profile.svg'. Thereafter, it will either execute the command inside the textbox (Command) or try to open the svg file with the default .svg reader (Default).",
			"Hint: All options under the 'Span' tab are based on the following reference: ¹Kozuch, S. & Shaik, S. Acc. Chem. Res. 2011, 44, 101."
			]

def is_str_float(i):
	try: float(i); return True
	except ValueError: return False

class Note(ttk.Notebook):
	def __init__(self,parent,*args,**kwargs):
		ttk.Notebook.__init__(self,parent,*args,**kwargs)
		for a,b in zip(pref.menu_g,pref.menu_h):
			setattr(self, a,TabFramePaths(self,name=b))
		self.tab_comparers = TabFrameComparers(self, name="Comparers")
		self.tab_connections = TabFrameConnections(self,name="Connections")
		self.pack(fill="both", expand=True)

class ScrollingFrame(tk.Frame):
	# Adapted from stackoverflow.com/a/3092341/13702856
	def __init__(self, parent):
		tk.Frame.__init__(self, parent)
		self.canvas = tk.Canvas(self)
		self.frame = tk.Frame(self.canvas)
		self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
		self.canvas.configure(yscrollcommand=self.vsb.set)
		self.vsb.pack(side="right", fill="y")
		self.canvas.pack(side="left", fill="both", expand=True)
		self.canvas.create_window((4,4), window=self.frame, anchor="nw", tags="self.frame")
		self.frame.bind("<Configure>", self.onFrameConfigure)
		self.pack(side="bottom")
	def onFrameConfigure(self, event):
		'''Reset the scroll region to encompass the inner frame'''
		self.canvas.configure(scrollregion=self.canvas.bbox("all"))

class TabFramePaths(ttk.Frame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.Frame.__init__(self,parent,*args,**kwargs)
		self.parent = parent
		self.parent.add(self, text=name)
		g = tk.Frame(self)
		z=tk.Frame(g)
		z.pack(side="left",fill="x")
		g.pack(side="top",fill="x")
		y = ScrollingFrame(self)
		y.pack(side="bottom",fill="both",expand=True)
		x = y.frame
		#########
		self._build_options(z)
		#########
		self.widths = [3,3,16,10,10,1,1,6]
		for a,b,c in zip(range(len(self.widths)),[" ","Type",'Structure Name','Free Energy','Enthalpy',"Move","",'Alignment'],self.widths):
			if a == 6: continue
			label = tk.Label(x, text=b,width=4 if a == 5 else c)
			if a == 5: label.grid(column=a, row=9, rowspan=1, columnspan = 2,sticky="news")
			else: label.grid(column=a, row=9, rowspan=1,sticky="news")
		#########
		self.data = [[None,None,None,None,None] for _ in range(pref.n_structures)]
		#########
		for n in range(pref.n_structures):
			label = tk.Label(x, text='#{}'.format(n+1),width=self.widths[0])
			label.grid(column=0, row=10+n, rowspan=1)
			for b in [1,2,3]:
				self.data[n][b] = tk.Entry(x,justify=tk.CENTER,bd=2,width=self.widths[1+b])
				self.data[n][b].insert(0,"")
				self.data[n][b].grid(column=1+b, row=10+n,padx="0",sticky="news")
			if not n == 0:
				button = tk.Button(x, text=u'\u2191', command=lambda x = n: self._move(x,x-1), padx="1")
				button.config(width=self.widths[5])
				button.grid(column=5, row=10+n)
			if not n+1 == pref.n_structures:
				button = tk.Button(x, text=u'\u2193', command=lambda x = n: self._move(x,x+1), padx="1")
				button.config(width=self.widths[6])
				button.grid(column=6, row=10+n)
			self.data[n][0] = tk.StringVar()
			menu = tk.OptionMenu(x,self.data[n][0],*pref.menu_z)
			self.data[n][0].set(pref.menu_z[0])
			menu.config(width=self.widths[1])
			menu.grid(column=1, row=10+n)
			self.data[n][4] = tk.StringVar()
			menu = tk.OptionMenu(x,self.data[n][4],*pref.menu_d)
			self.data[n][4].set(pref.menu_d[1])
			menu.config(width=self.widths[7])
			menu.grid(column=7, row=10+n)
	def _build_options(self,x):
		self.option_menu = ttk.Frame(x)
		setattr(self.option_menu,"line_opt_data",[[None,None,None],[None,None,None]])
		for n,name in enumerate(["Main[Color/Width/Strike]","Link[Color/Width/Strike]"]):
			box = self.boxify(self.option_menu,name=name,column=n)
			for a,b in zip([0,1,2],[pref.menu_a,pref.menu_b,pref.menu_c]):
				self.option_menu.line_opt_data[n][a] = tk.StringVar()
				color_menu = tk.OptionMenu(box, self.option_menu.line_opt_data[n][a],*b)
				self.option_menu.line_opt_data[n][a].set(b[[1,1,0][a] if n == 0 else [1,0,2][a]])
				color_menu.config(width=["9","1","7"][a])
				color_menu.grid(column=a,row=0,sticky="news")
		self.option_menu.grid(column=0,row=0,columnspan = 8,rowspan=8,sticky="news")
	def _move(self,n,x):
		line_n = [a.get() for a in self.data[n]]
		other = [a.get() for a in self.data[x]]
		for i, a in enumerate(other):
			if type(self.data[n][i]) is tk.Entry:
				self.data[n][i].delete(0,tk.END)
				self.data[n][i].insert(0,a)
			elif type(self.data[n][i]) is tk.StringVar:
				self.data[n][i].set(a)
		for i, a in enumerate(line_n):
			if type(self.data[x][i]) is tk.Entry:
				self.data[x][i].delete(0,tk.END)
				self.data[x][i].insert(0,a)
			elif type(self.data[x][i]) is tk.StringVar:
				self.data[x][i].set(a)
	@staticmethod
	def boxify(other, name, column):
		box = ttk.LabelFrame(other, text=name)
		box.grid(column=column, row=0, sticky="news")
		return box
class TabFrameConnections(ttk.Frame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.Frame.__init__(self,parent,*args,**kwargs)
		self.parent = parent
		self.parent.add(self, text=name)
		self.data = [[None,None,None,None,None,None,None] for _ in range(pref.n_connectors)]
		for n in range(pref.n_connectors):
			con = tk.LabelFrame(self,text="Conector {}".format(n+1))
			con.grid(column=0, row=n*2,  columnspan=5, pady="0",padx="2", rowspan=2,sticky="news")
			for b in range(2):
				label = tk.Label(con, text="From path" if b ==0 else "to path")
				label.grid(column=b*4+0, row=0, sticky="w")
				self.data[n][b*2] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2],"",*pref.menu_e)
				menu.config(width="2")
				menu.grid(column=b*4+1, row=0,sticky = "e")
				label = tk.Label(con, text=", number" if b ==0 else ", number")
				label.grid(column=b * 4 + 2, row=0, sticky="w")
				self.data[n][b*2+1] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2+1],"",*list(range(1,pref.n_structures+1)))
				menu.config(width="2")
				menu.grid(column=b*4+3, row=0,sticky = "e")
			label = tk.Label(con,text="Color/Width/Strike:")
			label.grid(column=0,row=1,columnspan=2)
			for a, b in zip([0, 1, 2], [pref.menu_a, pref.menu_b, pref.menu_c]):
				self.data[n][a+4] = tk.StringVar()
				self.color_menu = tk.OptionMenu(con, self.data[n][a+4], *b)
				self.data[n][a+4].set(b[0])
				self.color_menu.config(width="12")
				self.color_menu.grid(column=2 + a*2,columnspan =2, row=1)
class TabFrameComparers(ttk.Frame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.Frame.__init__(self,parent,*args,**kwargs)
		self.parent = parent
		self.parent.add(self, text=name)
		self.data = [[None for _ in range(10)] for _ in range(pref.n_connectors)]
		for n in range(pref.n_comparers):
			con = tk.LabelFrame(self,text="Comparer {}".format(n+1))
			con.grid(column=0, row=n*2,  columnspan=5, pady="0",padx="2", rowspan=2,sticky="news")
			for b in range(2):
				label = tk.Label(con, text="Path:" if b ==0 else ", vs. path:")
				label.grid(column=b*4+0, row=0, sticky="w")
				self.data[n][b*2] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2],"",*pref.menu_e)
				menu.config(width="2")
				menu.grid(column=b*4+1, row=0,sticky = "e")
				label = tk.Label(con, text=", number" if b ==0 else ", number")
				label.grid(column=b * 4 + 2, row=0, sticky="w")
				self.data[n][b*2+1] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2+1],"",*list(range(1,pref.n_structures+1)))
				menu.config(width="2")
				menu.grid(column=b*4+3, row=0,sticky = "e")
			label = tk.Label(con,text="Color/Width/Strike:")
			label.grid(column=0,row=1,columnspan=2)
			for a, b in zip([0, 1, 2], [pref.menu_a, ["1","1.5","2","3"], pref.menu_c]):
				self.data[n][a+4] = tk.StringVar()
				self.color_menu = tk.OptionMenu(con, self.data[n][a+4], *b)
				self.data[n][a+4].set(b[0])
				self.color_menu.config(width="12")
				self.color_menu.grid(column=2 + a*2,columnspan =2, row=1)
			self.data[n][4].set("green")
			self.data[n][5].set("1.5")
			self.data[n][6].set("dashed3")
			label = tk.Label(con,text="Vertical/Horizontal/Label:")
			label.grid(column=0,row=2,columnspan=2)
			x = ["normal","reverse"]
			y = ["left","right","midle","p_left","xp_left","p_right","xp_right","average"]
			z = ["left","right","fliped_left","fliped_right"]
			for a, b in zip([0, 1, 2], [x, y, z]):
				self.data[n][a+4+3] = tk.StringVar()
				self.color_menu = tk.OptionMenu(con, self.data[n][a+4+3], *b)
				self.data[n][a+4+3].set(b[{0:0,1:6,2:0}[a]])
				self.color_menu.config(width="12")
				self.color_menu.grid(column=2 + a*2,columnspan =2, row=2)

class GeneralMenu(tk.LabelFrame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.LabelFrame.__init__(self,parent,text=name,*args,**kwargs)
		self.titles = ["Main Title", "ΔE in kcal·mol⁻¹", "Reaction coordinate"]
		self.main = []
		self.span = []
		self.style = []
		self.horizontal = []
		self.vertical = []
		self.labels = []
		self.plot = []
		self.command = ""
		self.aesthetics = []
		###BUILD
		self._build_all()
		if hasattr(pref,"filename"):
			self.f = pref.filename
			self.load_state(getattr(pref,"filename"))
	def _build_all(self):
		self.note = ttk.Notebook(self.boxify("Advanced options", 2))
		self.note.grid(column=0, row=0, sticky="news")
		self.note.grid_columnconfigure(0, weight=1)
		self._change_win_title("Unsaved")
		self._build_main_opt()
		self._build_span_opt()
		self._build_style_sel()
		self._buid_horizontal_sel()
		self._buid_vertical_sel()
		self._buid_labels_sel()
		self._build_plot_sel()

		self._build_titles(6)
		self._build_loadsave(7)
		if pref.trickster: self._build_generator(8)
		self._build_preview(9)
		self._build_message(10)

	def _build_main_opt(self):
		box = self.framefy("Main")
		options = ["Plot and use enthalpy instead of free energy","Use comma as decimal separator","Include complementary data (H or G values)"]
		for i,a in enumerate(options):
			self.main.append(tk.IntVar(value=0))
			c1 = tk.Checkbutton(box, text=a, variable=self.main[i], onvalue=1, offvalue=0)
			c1.grid(column=0,row=i,sticky="w")
		for n in range(3):
			box.grid_rowconfigure(n, weight=1)
	def _build_span_opt(self):
		box = self.framefy("Span")
		box.grid_columnconfigure(1, weight=1)
		options = ["Atempt span calculation","Ignore structure type (TS/INT)","Use big indicator arrows"]
		for i,a in enumerate(options):
			self.span.append(tk.IntVar(value=0))
			c1 = tk.Checkbutton(box, text=a, variable=self.span[i], onvalue=1, offvalue=0)
			c1.grid(column=0,row=i,columnspan=1,sticky="w")
		label = tk.Label(box, text="Input units:")
		label.grid(column=1, row=1,sticky="news")
		self.span.append(tk.StringVar())
		options = ["kcal/mol","kJ/mol","eV"]
		menu = tk.OptionMenu(box, self.span[-1], *options)
		self.span[-1].set(options[0])
		menu.config(width="8")
		menu.grid(column=3, row=1,sticky="w")
		label = tk.Label(box, text="Temperature (°C):")
		label.grid(column=1, row=2, columnspan = 2)
		self.span.append(tk.Entry(box, justify=tk.CENTER, bd=3, width=6))
		self.span[-1].insert(0, "25")
		self.span[-1].grid(column=3, row=2, padx="3",pady="4", sticky="news")
		for n in range(3):
			box.grid_rowconfigure(n, weight=1)
	def _build_style_sel(self):
		options = list(pref.image.keys())
		box = self.framefy("Graph Style")
		label = tk.Label(box, text=">-- Graph Style preview -->")
		label.grid(column=0, row=0, columnspan=2, sticky="news")

		self.preview = tk.Message(box, text=pref.image[options[0]], font="TkFixedFont", relief = tk.RIDGE)
		self.preview.grid(column=3, row=0, rowspan = 3, sticky="e")

		label = tk.Label(box, text="Graph Style:")
		label.grid(column=0, row=1, sticky="w")
		self.style.append(tk.StringVar())
		self.style[-1].set(options[0])

		update = lambda x: self.preview.configure(text=pref.image[x])
		menu = tk.OptionMenu(box, self.style[-1], *options, command=update)
		menu.config(width="10")
		menu.grid(column=1, row=1,sticky="w")

		options = ["0","1","2"]
		label = tk.Label(box, text="Grid decimal places:")
		label.grid(column=0, row=2, sticky="w")
		self.style.append(tk.StringVar())
		self.style[-1].set(options[1])

		menu = tk.OptionMenu(box, self.style[-1], *options)
		menu.config(width="10")
		menu.grid(column=1, row=2, sticky="w")

		box.grid_columnconfigure(2, weight=1)
		for n in range(3):
			box.grid_rowconfigure(n, weight=1)
	def _buid_horizontal_sel(self):
		box = self.framefy("Horizontal")
		a = [2 * n for n in range(11)]
		b = [10 * n for n in range(11)]
		c = [5 * n + 60 for n in range(11)]
		options = [a,b,c]
		for i, x , y in zip(range(3), [["Plateau width:"], ["Starting offset (X):","Ending offset (X)"], ["Plateau spacing:"]], options ):
			for idx,z in enumerate(x):
				label = tk.Label(box, text=z)
				label.grid(column=idx*2, row=i, sticky="w")
				self.horizontal.append(tk.StringVar())
				menu = tk.OptionMenu(box, self.horizontal[-1], *y)
				menu.config(width="8")
				menu.grid(column=1+idx*2, row=i, sticky="e")
		box.grid_columnconfigure(3, weight=1, minsize="50")
		self.horizontal[0].set(a[5])
		self.horizontal[1].set(b[5])
		self.horizontal[2].set(b[5])
		self.horizontal[3].set(c[5])
	def _buid_vertical_sel(self):
		box = self.framefy("Vertical")
		a = [5 * n for n in range(11)]
		b = [5 * n for n in range(11)]
		options = [a,b]
		for i, x , y in zip(range(2), ["Top offset (Y)", "Bottom offset (Y)"], options):
			label = tk.Label(box, text=x)
			label.grid(column=0, row=i, sticky="w")
			self.vertical.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.vertical[i], *y)
			menu.config(width="8")
			menu.grid(column=1, row=i, sticky="e")
		box.grid_columnconfigure(3, weight=1, minsize="150")
		self.vertical[0].set(a[2])
		self.vertical[1].set(a[0])
	def _buid_labels_sel(self):
		box = self.framefy("Labels")
		a = ["   ", "( )", "[ ]", r"{ }", '" "', "' '"]
		for i,x in enumerate(["G:", "H:"]):
			label = tk.Label(box, text=x)
			label.grid(column=2 * i, row=0,sticky="w")
			self.labels.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.labels[i], *a)
			menu.config(width="8")
			menu.grid(column=i * 2 + 1, row=0, sticky="e")
		self.labels[0].set(a[0])
		self.labels[1].set(a[2])

		#Decimal
		label = tk.Label(box, text="Decimal places:")
		label.grid(column=0, row=1,columnspan=2,sticky="w")
		self.labels.append(tk.StringVar())
		menu = tk.OptionMenu(box, self.labels[-1], *["0","1","2"])
		menu.config(width="8")
		menu.grid(column=2, row=1,columnspan=2,sticky="e")
		self.labels[-1].set("1")

		#TS MARK
		label = tk.Label(box, text="TS Mark:")
		label.grid(column=0, row=2,columnspan=2,sticky="w")
		self.labels.append(tk.StringVar())
		menu = tk.OptionMenu(box, self.labels[-1], *[" ", "‡ (big)", "‡ (small)"])
		menu.config(width="8")
		menu.grid(column=2, row=2,columnspan=2,sticky="e")
		box.grid_columnconfigure(5, weight=1,minsize="180")
		self.labels[-1].set(" ")

		#ADJUST GRID
		for n in range(3):
			box.grid_rowconfigure(n, weight=1)
		for n in range(4):
			box.grid_columnconfigure(n, weight=1)

	def _build_aesthetics(self):
		box = self.framefy("Aesthetics")
		a = ["   ", "( )", "[ ]", r"{ }", '" "', "' '"]
		b = [a * 2 for a in range(11)]
		c = [1,2]
		e = [a,a,b,c]
		f = ["G:","H:","Width:","Decimal"]
		for a,(b,c) in enumerate(zip(f,e)):
			label = tk.Label(box,text=b)
			label.grid(column=2*a,row=0)
			self.aesthetics.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.aesthetics[a], *c)
			menu.config(width="2")
			menu.grid(column=a * 2 + 1, row=0)
		for i,a in enumerate(["   ","( )",10,1]):
			self.aesthetics[i].set(a)
		a = [0 + a * 10 for a in range(11)]
		b = [60 + a * 5 for a in range(11)]
		c = [" ","‡ (big)","‡ (small)"]
		d = ["X offset:","X dist:","TS mark:"]
		e = [a,b,c]
		for a, (b, c) in enumerate(zip(d, e)):
			label = tk.Label(box, text=b)
			label.grid(column=0 if a ==0 else 2 * a +1, row=1, columnspan = 2 if a ==0 else 1)
			self.aesthetics.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.aesthetics[a+4], *c)
			menu.config(width="8" if a == 2 else "2")
			menu.grid(column=a * 2 + 2, row=1, columnspan = 3 if a == 2 else 1, sticky="news" if a==2 else "")
		for i, a in enumerate([40, 80, " "]):
			self.aesthetics[i+4].set(a)
	def _build_plot_sel(self):
		box = self.framefy("Plot")
		for i,a in enumerate(pref.menu_h):
			self.plot.append(tk.IntVar(value=1))
			c1 = tk.Checkbutton(box, text=a, variable=self.plot[i], onvalue=1, offvalue=0)
			c1.grid(column=i%4,row=i//4, sticky="w")
		n = len(pref.menu_h)
		self.plot.append(tk.IntVar(value=1))
		c1 = tk.Checkbutton(box, text="Comparers", variable=self.plot[n], onvalue=1, offvalue=0)
		c1.grid(column=0,row=((n-1)//4)+1,columnspan = 2, sticky="w")
		self.plot.append(tk.IntVar(value=1))
		c1 = tk.Checkbutton(box, text="Connections", variable=self.plot[n+1], onvalue=1, offvalue=0)
		c1.grid(column=2,row=((n-1)//4)+1,columnspan = 2, sticky="w")
		for n in range(4):
			box.grid_columnconfigure(n, weight=1)
		for n in range(3):
			box.grid_rowconfigure(n, weight=1)
	def _build_titles(self,idx):
		box = self.boxify("Titles",idx)
		for a,b,c in zip([0,1,2],self.titles,["Main:","y:","x:"]):
			label = tk.Label(box, text=c,width=10)
			label.grid(column=0, row=a)
			self.titles[a] = tk.Entry(box, justify=tk.CENTER, bd=2, width=50)
			self.titles[a].insert(0, b)
			self.titles[a].grid(column=1, row=a, padx="0", sticky="news")
	def _build_loadsave(self,idx):
		box = self.boxify("Close, Load & Save Data", idx)
		label = tk.Label(box, text="Path's and connection's info")
		label.grid(column=0,row=0,sticky="w")
		button = tk.Button(box, text="Close", command=self._blank_state, padx="1")
		button.config(width=7)
		button.grid(column=1, row=0, sticky="e")
		button = tk.Button(box, text="Load", command=self.load_state, padx="1")
		button.config(width=7)
		button.grid(column=2,row=0,sticky="e")
		button = tk.Button(box, text="Save as", command=self._save_as, padx="1")
		button.config(width=7)
		button.grid(column=3,row=0,sticky="e")
		button = tk.Button(box, text="Save", command=self._save, padx="1")
		button.config(width=7)
		button.grid(column=4,row=0,sticky="e")
		box.grid_columnconfigure(0, weight=1)
	def _build_generator(self,idx):
		box = self.boxify("Generate random PES", idx)
		label = tk.Label(box, text="Random PES generator")
		label.pack(side=tk.LEFT)
		button = tk.Button(box, text="Fill in data", command=self._ask_confirmation, padx="1")
		button.config(width=10)
		button.pack(side=tk.RIGHT)
	def _build_preview(self,idx):
		box = self.boxify("Preview with either command or default manager, or save svg file", idx)
		self.command = tk.Entry(box, justify=tk.CENTER, bd=4)
		self.command.insert(0, pref.command_line)
		self.command.grid(column=0,row=0,sticky="news")
		button = tk.Button(box, text="Command", command=self.run_data_a, padx="1")
		button.config(width=8)
		button.grid(column=1,row=0,sticky="e")
		button = tk.Button(box, text="Default", command=self.run_data_b, padx="1")
		button.config(width=8)
		button.grid(column=2,row=0,sticky="e")
		button = tk.Button(box, text="Save svg", command=self.return_svg, padx="1")
		button.config(width=8)
		button.grid(column=3,row=0,sticky="e")
		box.grid_columnconfigure(0, weight=1)
	def _build_message(self,idx):
		box = self.boxify("Message",idx)
		m = tk.Message(box)
		scrollbar = tk.Scrollbar(box)
		scrollbar.grid(column=1,row=0,sticky="nes")
		self.msg = tk.Text(box,
						   width=40,
						   yscrollcommand=scrollbar.set,
						   height=12 if pref.trickster else 18,
						   state="disabled",
						   background=m.cget("background"),
						   relief="flat",
						   wrap=tk.WORD,
						   font=('TkFixedFont',8))
		m.destroy()
		self.msg.grid(column=0,row=0,sticky="news")
		if not pref.xlsx:
			self.message("Welcome!\n\nTo enable .xlsx file support, please install openpyxl python library via the shell command:\npython3 -m pip install openpyxl")
		else:
			self.message("Welcome!\n\n{}".format(random.choice(pref.hints)))
		box.grid_rowconfigure(0, weight=1)
		box.grid_columnconfigure(0, weight=1)
		self.grid_rowconfigure(idx, weight=1)
	def _save(self):
		try:
			if self.f.endswith(".xlsx") and pref.xlsx:
				from openpyxl import Workbook
				wb = Workbook()
				wb.remove(wb.active)
				for a,b in zip(self.gen_data(type=".xlsx"),pref.menu_e):
					sheet = wb.create_sheet(title=f'Path {b}')
					for i,c in enumerate(a,start=1):
						sheet.append(c[1:4])
				try:
					wb.save(self.f)
				except PermissionError:
					self.message(f"Error while saving file!\nIs the file:\n'{self.f}' already open?")
			else:
				try:
					with open(self.f,"w") as out:
						if self.f.endswith(".txt"):
							txt = "\n".join(a for a in self.gen_data(type=".txt") if len(a.split()) >= 1)
							out.write(txt)
				except PermissionError:
					self.message(f"Error while saving file!\nIs the file:\n'{self.f}' already open?")
		except AttributeError: self._save_as()
		except FileNotFoundError: self._save_as()
	def _save_as(self):
		self.f = tk.filedialog.asksaveasfilename(**pref.allowed_extensions)
		self._change_win_title(self.f)
		if any(self.f.endswith(a) for a in (".sff",".txt",".xlsx")):self._save()

	def load_state(self,file_n=None):
		if file_n is None:
			file_n = tk.filedialog.askopenfilename(**pref.allowed_extensions)
		try:
			if file_n.endswith(".xlsx") and pref.xlsx:
				self._blank_state(ask=False)
				import openpyxl
				try:
					wb = openpyxl.load_workbook(file_n)
				except:
					self.message(f"Could not read {file_n} as xlsx file!\nAre you sure this is a proper xlsx file?")
					return
				notes = [getattr(note, a) for a in pref.menu_g]
				exceeded = False
				for a,b in zip(wb.sheetnames, notes):
					sheet = wb[a]
					for n in range(1,pref.n_structures+10):
						if n > pref.n_structures:
							if any(sheet.cell(row=n,column=i).value is None for i in range(1,4)):
								exceeded = True
							continue
						for i in range(1,4):
							if sheet.cell(row=n,column=i).value is None: continue
							b.data[n-1][i].insert(0,str(sheet.cell(row=n,column=i).value))
				if exceeded:
					self.message("Exceeding number of structures")
			elif file_n.endswith(".txt"):
				with open(file_n, mode="r") as file:
					self._blank_state(ask=False)
					all_tabs = {}
					tab_data = []
					for line in file.read().splitlines():
						line = line.split()
						non_hash = True if len(line) == 1 and not line[0].startswith("#") else False
						if len(line) >= 2 or non_hash:
							tab_data.append(line)
						elif len(line) == 1 and any(line[0] == f"#{a}" for a in pref.menu_e):
							all_tabs[line[0]] = tab_data
							tab_data = []
					if len(all_tabs) == 0 and len(tab_data) != 0:
						all_tabs["#A"] = tab_data
					missing = [b for b in [f"#{a}" for a in pref.menu_e] if b not in all_tabs.keys()]
					for a in missing: all_tabs[a] = []
					notes = [getattr(note, a) for a in pref.menu_g]
					exceeded = False
					for a,b in zip(notes,sorted(all_tabs.keys())):
						for i,c in enumerate(all_tabs[b]):
							if i >= pref.n_structures:
								exceeded = True
								continue
							for n in range(3):
								try: a.data[i][n+1].insert(0,c[n])
								except IndexError: pass
					if exceeded:
						self.message("Exceeding number of structures")
			else:
				self.message(f"Unrecognized file {file_n}")
		except FileNotFoundError:
			self.message("File not found!")
			return
		finally:
			self._change_win_title(file_n)
			self.f = file_n

	def fill_in(self):
		size = random.random()+0.5
		max_value = min(len(pref.alphabet), pref.n_structures)
		lenght = random.randint(5,12)
		tab = getattr(note,[a for a in pref.menu_g][note.index(note.select())])
		for i,n in zip(range(max_value),pref.alphabet):
			value = size*random.randrange(-20,20)-i*size*2
			for idx in range(len(tab.data[i])):
				if idx == 1:
					tab.data[i][idx].delete(0, tk.END)
					if i+1 < lenght: tab.data[i][idx].insert(0, n)
					elif i + 1 == lenght: tab.data[i][idx].insert(0, "A'")
				elif idx == 2:
					tab.data[i][idx].delete(0, tk.END)
					if i < lenght: tab.data[i][idx].insert(0,"{:.2f}".format(value))
				elif idx == 3:
					tab.data[i][idx].delete(0, tk.END)
					if i < lenght: tab.data[i][idx].insert(0,"{:.2f}".format(value + random.choice([-random.random(), +random.random()])))
				elif idx == 4:
					tab.data[i][idx].set(pref.menu_d[1])
		max_v, min_v = None, None
		for i in range(max_value):
			if max_v is None: max_v = [i,tab.data[i][2].get()]
			if min_v is None: min_v = [i,tab.data[i][2].get()]
			if i < lenght and float(tab.data[i][2].get()) > float(max_v[1]): max_v = [i, tab.data[i][2].get()]
			if i < lenght and float(tab.data[i][2].get()) < float(min_v[1]) : min_v = [i, tab.data[i][2].get()]
			if i == 0: tab.data[i][0].set("INT")
			elif i+1 == lenght: tab.data[i][0].set("INT")
			elif i >= lenght: tab.data[i][0].set("  ")
			else:
				if float(tab.data[i-1][2].get()) < float(tab.data[i][2].get()) > float(tab.data[i+1][2].get()):
					tab.data[i][0].set("TS")
				else:
					tab.data[i][0].set("INT")
	def _ask_confirmation(self):
		if note.index(tk.END) -2 <= note.index(note.select()):
			self.message("Cannot fill in data for connection and comparer tabs!\n")
			return
		msgbox = tk.messagebox.askquestion(
			f'Fill in random PES cycle at {pref.menu_h[note.index(note.select())]}',
			'Are you sure? All unsaved data will be lost!', icon='warning')
		if msgbox == "yes":
			self.fill_in()
			self._change_win_title("Unsaved")
			if hasattr(self,"f"): del(self.f)
	def _change_win_title(self,path):
		window.title(f"{pref.name} @ {path}")
	def _blank_state(self,ask=True):
		if ask:
			msgbox = tk.messagebox.askquestion('Close document', 'Are you sure? All unsaved data will be lost!', icon='warning')
			if msgbox != "yes":return
			self._change_win_title("Unsaved")
			if hasattr(self,"f"): del(self.f)
		for a in [getattr(note,a) for a in pref.menu_g]:
			for i in range(pref.n_structures):
				for idx in range(5):
					if idx == 0: a.data[i][idx].set(pref.menu_z[0])
					if idx in [1,2,3]: a.data[i][idx].delete(0, tk.END)
					if idx  == 4: a.data[i][idx].set(pref.menu_d[1])
		for a in [getattr(note,a) for a in pref.menu_g]:
			for n in range(2):
				for idx, b in zip([0, 1, 2], [pref.menu_a, pref.menu_b, pref.menu_c]):
					a.option_menu.line_opt_data[n][idx].set(b[[1,1,0][idx] if n == 0 else [1,0,2][idx]])
		for a in range(pref.n_connectors):
			for b in range(4):
				note.tab_connections.data[a][b].set("")
			for b,c in zip(range(3),[pref.menu_a, pref.menu_b, pref.menu_c]):
				note.tab_connections.data[a][b+4].set(c[0])
	def message(self,text):
		now = datetime.datetime.now()
		self.msg.configure(state="normal")
		self.msg.tag_add("start", "0.0", tk.END)
		self.msg.tag_config("start", foreground="grey")
		if type(text) == str: text = [text]
		for txt in text:
			self.msg.insert("1.0",txt+"\n")
		self.msg.insert("1.0", "[" + ":".join(["{:02d}".format(a) for a in [now.hour, now.minute, now.second]]) + "] "+"\n")
		self.msg.configure(state="disabled")
	def boxify(self,name,row):
		box = ttk.LabelFrame(self, text=name)
		box.grid(column=0, row=row, sticky="news")
		box.grid_columnconfigure(0, weight=1)
		return box
	def framefy(self,name):
		x = tk.Frame()
		x.grid(column=0, row=0, sticky="news")
		x.grid_columnconfigure(0, weight=1)
		self.note.add(x,text=name)
		return x
	def print_data(self):
		notes = [getattr(note,a) for a in pref.menu_g]
		for a,b in zip(notes,pref.menu_e):
			print(f"NOTE {b}")
			for idx,line in enumerate(getattr(a,"data")):
				if any(c.get().strip() != "" for c in line[:-1]):
					print(f"#{idx+1}",[n.get() for n in line])
		print("NOTE CONNECTIONS")
		for idx,a in enumerate(note.tab_connections.data):
			if any(c.get().strip() != "" for c in a[:-3]):
				print(f"#{idx+1}", [n.get() for n in a])
	def gen_data(self,type):
		notes = [getattr(note,a) for a in pref.menu_g]
		txt_data = []
		xlsx_data = []
		for a,b in zip(notes,pref.menu_e):
			xlsx = []
			for idx,line in enumerate(getattr(a,"data")):
				c = [n.get() for n in line]
				txt_data.append("{:<20} {:>10} {:>10}".format(*c[1:4]))
				xlsx.append(c)
			txt_data.append("#{}".format(b))
			xlsx_data.append(xlsx)
		if type == ".txt":
			return txt_data
		elif type == ".xlsx":
			return xlsx_data
	def save_svg_as(self):
		return tk.filedialog.asksaveasfilename(defaultextension=".svg", title="Save svg", filetypes=[("Scalable Vector Graphics", ".svg")])
	def run_data_a(self):
		self.return_svg(promp=False); os.system(self.command.get())
	def run_data_b(self):
		self.return_svg(promp=False)
		filename = os.path.join(os.getcwd(), ".E_profile.svg")
		if sys.platform == "win32" or os.name == "nt":
			os.startfile(filename)
		else:
			opener = "open" if sys.platform == "darwin" else "xdg-open"
			subprocess.call([opener, filename])

	def return_svg(self,promp=True):
		svg_name = None if promp == False else self.save_svg_as()
		msg = SvgGenEsp(self)
		msg = msg.save_svg(svg_name)
		if not msg is None: self.message(msg)

class SvgGenEsp:
	def __init__(self,options):
		self.options = options # Please only use this at __init__

		#MAIN
		m_options = ["energy", "comma", "plot"]
		self.main = {a: b.get() for a, b in zip(m_options, getattr(self.options,"main"))}
		self.e_source = 4 if getattr(self.options,"main")[0].get() == 1 else 3
		self.e_complement = 3 if self.e_source == 4 else 4
		self.comma = True if self.main["comma"] == 1 else False
		self.plot_np = True if self.main["plot"] == 1 else False

		#SPAN
		s_options = ["span","irrespective","big_arrow","units","temperature"]
		self.span = {a:b.get() for a,b in zip(s_options,getattr(self.options,"span"))}
		self.span_worthy = True
		self.span_request = True if self.span["span"] == 1 else False
		self.big_arrow = True if self.span["big_arrow"] == 1 else False

		#GRAPHIC STYLE
		self.frame = getattr(self.options,"style")[0].get()
		self.grid_decimal = getattr(self.options, "style")[1].get()

		#HORIZONTAL
		self.wide = [int(getattr(self.options,"horizontal")[0].get()) * a + b for a,b in zip([-1,1],[20,40])]
		self.x_start_offset = int(getattr(self.options,"horizontal")[1].get())
		self.x_end_offset = int(getattr(self.options, "horizontal")[2].get())
		self.x_space = int(getattr(self.options,"horizontal")[3].get())

		#VERTICAL
		self.top_height = int(getattr(self.options,"vertical")[0].get())
		self.bottom_height = 400 - int(getattr(self.options,"vertical")[1].get())

		#LABELS
		self.g_h_labels = {a: getattr(self.options,"labels")[b].get() for a,b in zip(["g","h"],[0,1])}
		self.e_decimal = getattr(self.options,"labels")[2].get()
		self.ts_mark = getattr(self.options,"labels")[3].get()

		#PLOT
		self.plot = [a.get() for a in getattr(self.options, "plot")]
		self.plot_path = {a:bool(b) for a,b in zip(pref.menu_e,self.plot)}

		#TITLE
		self.main_title = getattr(self.options,"titles")[0].get()
		self.y_title = getattr(self.options,"titles")[1].get()
		self.x_title = getattr(self.options, "titles")[2].get()

		# RETURN
		self.svg_code = ['<?xml version="1.0" encoding="UTF-8" ?>']
		self.msg = []

		if self.span_request:
			self.temperature = self._verify_temp(self.span["temperature"])

		# CONECTORS
		self.conectors = [[b.get() for b in a] for a in note.tab_connections.data]
		# COMPARERS
		x = ("A","1","B","2","S1","S2","S3","S4","S5","S6")
		self.comparers = [{l:note.tab_comparers.data[n][i].get() for i,l in zip(range(10),x)} for n in range(pref.n_comparers)]



		# DATA OPTIONS
		fc = lambda a: getattr(note, "tab_{}".format(a.lower())).option_menu.line_opt_data
		self.path_options = {a: [[c.get() for c in b] for b in fc(a)] for a in pref.menu_e}
		# DATA
		dt = lambda a: enumerate(getattr(note,a).data)
		fa = lambda idx,c: float(c.get().replace(",",".")) if idx == self.e_source-1 else c.get()
		fb = lambda b: is_str_float(b[self.e_source-1].get().replace(",","."))
		self.raw_crt = [[[i+1,*[fa(idx,c) for idx,c in enumerate(b)]] for i,b in dt(a) if fb(b)] for a in pref.menu_g]
		self.raw_crt_dict = {a: b for a, b in zip(pref.menu_e, self.raw_crt) if self.plot_path[a] and b}
		self.paths = self.set_height()
		self.data_dict = {a: b for a, b in zip(pref.menu_e, self.paths)}
		self.plot_dict = {a:b for a,b in self.data_dict.items() if self.plot_path[a] and b}

	def _verify_temp(self,value):
		if is_str_float(value.replace(",",".")):
			if float(value) <= -273.15:
				self.span_worthy = False
				text = "Temperature should not be smaller than or equal to {} °C\n"
				self.msg.append(text.format(self.commafy("-273.15")))
				return None
			else:
				t = float(value) + 273.15
				text = "Temperature is set to {} K\n"
				self.msg.append(text.format(self.commafy("{:.2f}".format(t))))
				return t
		else:
			self.span_worthy = False
			text = "Unrecognized temperature: {}\n"
			self.msg.append(text.format(str(value)))
	def commafy(self,item):
		return str(item).replace(".", ",") if self.comma else str(item).replace(",", ".")
	@functools.lru_cache(maxsize=1)
	def max_value(self):
		return max(max(a[self.e_source] for a in path) for path in list(self.raw_crt_dict.values()) if path)
	@functools.lru_cache(maxsize=1)
	def min_value(self):
		return min(min(a[self.e_source] for a in path) for path in list(self.raw_crt_dict.values()) if path)
	@functools.lru_cache(maxsize=1)
	def delta_value(self):
		return self.max_value()-self.min_value()
	@functools.lru_cache(maxsize=1)
	def n_col(self):
		try: x = max(max(a[0] for a in path) for path in list(self.plot_dict.values()) if path)
		except ValueError: x = 0
		return x
	@functools.lru_cache(maxsize=1)
	def set_height(self):
		paths = []
		for idx_a, a in enumerate(self.raw_crt):
			path = []
			for idx_b, b in enumerate(a):  # For every structure
				try:
					height = 1 - (b[self.e_source] - self.min_value())/ self.delta_value()
					height = int(round(abs(height*(self.bottom_height - self.top_height) + self.top_height)))
				except ZeroDivisionError: height = int(self.top_height/2 + self.bottom_height/2)
				path.append([*b,height])
			paths.append(path)
		return paths
	def set_single_height(self,value):
		try:
			height = 1 - (value - self.min_value()) / self.delta_value()
			height = int(height * (self.bottom_height - self.top_height) + self.top_height +50)
		except ZeroDivisionError:
			height = int(self.top_height / 2 + self.bottom_height / 2)
		return height
	def set_horizontal(self,mult,add):
		value = (float(mult)-1)*int(self.x_space)
		value += int(self.x_start_offset) + int(add) + 80
		return int(value)
	def char_care(self,text):
		return str(text).encode("ascii", "xmlcharrefreplace").decode("utf-8")
	def is_avail(self,name,path_a,num_a,path_b,num_b,ignore_same=False):
		try:
			first = int(num_a) in [a[0] for a in self.data_dict[path_a]]
			second = int(num_b) in [a[0] for a in self.data_dict[path_b]]
			if int(num_a) == int(num_b) and path_a == path_b:
				text = f"{name}: Cannot conect items on same column and path"
				self.msg.append(text)
				return False
			elif int(num_a) == int(num_b) and not ignore_same:
				text = f"{name}: Cannot conect items on same column"
				self.msg.append(text)
				return False
			elif not self.plot_path[path_a]:
				return False
			elif not self.plot_path[path_b]:
				return False
			return first and second
		except KeyError:
			return False
		except ValueError:
			return False

	@functools.lru_cache(maxsize=1)
	def graph_frame(self):
		horizontal_end = self.set_horizontal(self.n_col(), self.x_end_offset)
		a = [
			'<svg width="{0}" viewBox="30 0 {0} 500" height="500" xmlns="http://www.w3.org/2000/svg">', #0 FRAME
			'    <line x1="100" y1="35" x2="100" y2="475" stroke="black" stroke-width="2"/>', #1 FULL VERTICAL LEFT LINE
			'    <line x1="100" y1="475" x2="{}" y2="475" stroke="black" stroke-width="2"/>', #2 FULL BOTTOM HORIZONTAL LINE
			'    <text x="{}" y="20" font-size="22" text-anchor="middle" fill="black">{}</text>', #3 MAIN TITLE
			'    <text x="-250" y="55" font-size="22" {} text-anchor="middle" fill="black">{}</text>', #4 Y AXIS TITLE
			'    <text x="{}" y="495" font-size="22" text-anchor="middle" fill="black">{}</text>', #5 X AXIS TITLE
			'    <polygon points="100,25 95,35 105,35" style="fill:black;stroke:black;stroke-width:1"/>', #6 TOP LEFT Y ARROW
			'    <polygon points="{0},475 {1},470 {1},480" style="fill:black;stroke:black;stroke-width:1"/>', #7 BOTTOM RIGHT X ARROW
			'    <polygon points="100,485 95,475 105,475" style="fill:black;stroke:black;stroke-width:1"/>', #8 BOTTOM RIGHT Y ARROW
			'    <line x1="150" y1="475" x2="{}" y2="475" stroke="black" stroke-width="2"/>', #9 PARTIAL BOTTOM HORIZONTAL LINE
			'    <polygon points="140,475 150,470 150,480" style="fill:black;stroke:black;stroke-width:1"/>',# 10 BOTTOM LEFT X ARROW (PARTIAL)
			'    <polygon points="90,475 100,470 100,480" style="fill:black;stroke:black;stroke-width:1"/>',# 11 BOTTOM LEFT X ARROW (FULL)
			'    <line x1="{0}" y1="35" x2="{0}" y2="475" stroke="black" stroke-width="2"/>',# 12 FULL VERTICAL RIGHT LINE
			'    <line x1="100" y1="35" x2="{}" y2="35" stroke="black" stroke-width="2"/>' #13 FULL TOP HORIZONTAL LINE
		]
		a[0] = a[0].format(horizontal_end + 75)
		a[2] = a[2].format(horizontal_end + 55)
		a[3] = a[3].format(self.set_horizontal(self.n_col()/2,55), self.char_care(self.main_title))
		a[4] = a[4].format('transform="rotate(-90)"', self.char_care(self.y_title))
		a[5] = a[5].format(self.set_horizontal(self.n_col()/2,55), self.char_care(self.x_title))
		a[6] = a[6]
		a[7] = a[7].format(horizontal_end + 65,horizontal_end + 55)
		a[8] = a[8]
		a[9] = a[9].format(horizontal_end + 55)
		a[10] = a[10]
		a[11] = a[11]
		a[12] = a[12].format(horizontal_end + 55)
		a[13] = a[13].format(horizontal_end + 55)

		code = {"Borderless":[a[n] for n in (0,3)],
				"Y only":[a[n] for n in (0,1,3,4,6,8)],
				"XY":[a[n] for n in (0,1,2,3,4,5,6,7)],
				"X only":[a[n] for n in (0,2,5,7,11)],
				"Frame":[a[n] for n in (0,1,2,3,4,5,12,13)],
				"X,Y":[a[n] for n in (0,1,3,4,5,6,7,8,9,10)]
				}

		self.svg_code.extend(code[self.frame])

	@functools.lru_cache(maxsize=1)
	def graph_grid(self):
		if self.frame == "Borderless" or self.frame == "X only": return
		if self.delta_value() == 0: return
		step_size = float("{:.0E}".format(abs(self.delta_value())/10))
		max_e = float("{:.0E}".format(abs(2*self.delta_value()+self.max_value())))
		if step_size == 0: return
		steps = [max_e-step_size*n for n in range(60)]
		for item in steps:
			value = int(self.set_single_height(item))
			if 50 < value < 465:
				item = {
					"0":"{:.0f}".format(item),
					"1":"{:.1f}".format(item),
					"2":"{:.02f}".format(item)
				}[self.grid_decimal]
				item = self.commafy(item)
				c = [
					'    <line x1="100" y1="{0}" x2="105" y2="{0}" stroke="black" stroke-width="2"/>',
					'    <text x="80" y="{}" text-anchor="middle" fill="black">{}</text>']
				c[0] = c[0].format(value)
				c[1] = c[1].format(value,item)
				self.svg_code.extend(c)

	@functools.lru_cache(maxsize=1)
	def graph_crt_points(self):
		for key,value in self.data_dict.items():
			if not self.plot_path[key]: continue
			opt_cri = self.path_options[key][0]
			opt_con = self.path_options[key][1]
			if not len(value) == len(set([a[0] for a in value])):
				text = "WARNING: Two or more structures are occupying the same block lane!"
				self.msg.append(text)
			l_c = [0, 0, 0]  # last collumn
			for idx, item in enumerate(value):
				c_p = [self.set_horizontal(item[0], self.wide[0]),
					   int(round(item[-1] + 50)),
					   self.set_horizontal(item[0], self.wide[1])]
				# [x1,y1,x2], y1=y2
				a = [
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}{}</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}</text>']
				x = pref.svg_repl[opt_cri[-1]]
				z = pref.placement[item[-2]][0 if self.plot_np else 1]
				trick_g = "g" if self.e_source == 3 else "h"
				trick_h = "h" if self.e_source == 3 else "g"
				item_rep = lambda x: float(item[x].replace(",",".") if type(item[x]) is str else item[x])
				digit_rounding = lambda x: {"0": "{:.0f}".format(item_rep(x)),"1": "{:.1f}".format(item_rep(x)),"2": "{:.2f}".format(item_rep(x))}[self.e_decimal]

				g = self.g_h_labels[trick_g][0] + self.commafy(digit_rounding(self.e_source)) + self.g_h_labels[trick_g][-1]
				h = self.g_h_labels[trick_h][0] + self.commafy(digit_rounding(self.e_complement) if is_str_float(item[self.e_complement].replace(",",".")) else item[self.e_complement]) + self.g_h_labels[trick_h][-1]
				ts_dict = {
				    " "        : "",
				    "‡ (big)":'<tspan dy="-7" font-family="arial" font-size=".7em">{}</tspan>'.format(self.char_care("‡")),
				    "‡ (small)":'<tspan dy="-7" font-family="monospace" font-size=".7em">{}</tspan>'.format(self.char_care("‡"))
				}
				ts_mark = ts_dict[self.ts_mark] if item[1] == "TS" else ""
				a[0] = a[0].format(c_p[0], c_p[1], c_p[2], c_p[1], opt_cri[0],opt_cri[1],x)
				a[1] = a[1].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[0], opt_cri[0],self.char_care(item[2]),ts_mark)
				a[2] = a[2].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[1],opt_cri[0],self.char_care(g))
				a[3] = a[3].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[2],opt_cri[0],self.char_care(h))
				self.svg_code.extend(a if self.plot_np else a[:-1])
				if not idx == 0:
					b = '    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>'
					x = pref.svg_repl[opt_con[-1]]
					b = b.format(l_c[2], l_c[1], c_p[0], c_p[1], opt_con[0],opt_con[1],x)
					self.svg_code.append(b)
				l_c = c_p
	@functools.lru_cache(maxsize=1)
	def graph_connectors(self):
		if not self.plot[-1] == 1: return
		for idx,i in enumerate(self.conectors):
			if not self.is_avail(f"Connector {idx+1}",*i[0:4]):continue
			i[1] = int(i[1])
			i[3] = int(i[3])
			start = next(n for n in self.data_dict[i[0]] if n[0] == i[1])
			end = next(n for n in self.data_dict[i[2]] if n[0] == i[3])
			con = [start,end]
			if con[0][0] > con[1][0]: con.reverse()
			if con[0][0] < con[1][0]:
				x = pref.svg_repl[i[6]]
				a = '    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>'
				a = a.format(self.set_horizontal(con[0][0], self.wide[1]),
							 con[0][-1] + 50,
							 self.set_horizontal(con[1][0], self.wide[0]),
							 con[1][-1] + 50, i[4], i[5],x)
				self.svg_code.append(a)
	@functools.lru_cache(maxsize=1)
	def graph_comparers(self):
		if not self.plot[-2] == 1: return
		for i,a in enumerate(self.comparers):
			if not self.is_avail(f"Comparer {i+1}",a["A"],a["1"],a["B"],a["2"],ignore_same=True): continue
			start = next(n for n in self.data_dict[a["A"]] if n[0] == int(a["1"]))
			end = next(n for n in self.data_dict[a["B"]] if n[0] == int(a["2"]))
			com = [start,end]
			ordered = com[0][0] < com[1][0]
			if not ordered:
				com = [end,start]
			if a["S4"] == "reverse":
				com.reverse()
			y1 = com[0][6] + 50
			y2 = com[1][6] + 50
			color = a["S1"]
			width = a["S2"]
			dashed = pref.svg_repl[a["S3"]]
			excess_y = 8 if com[0][-1] < com[1][-1] else -8
			excess_yb = 2 if com[0][-1] < com[1][-1] else -2
			text_pos = {
				"left":[-5,"-90"],
				"right":[15, "-90"],
				"fliped_left": [-15, "90"],
				"fliped_right": [5, "90"]
			}[a["S6"]]
			ordered = com[0][0] < com[1][0]
			digit_rounding = {"0": "{:.0f}", "1": "{:.1f}", "2": "{:.2f}"}[self.e_decimal]
			label = self.commafy(digit_rounding.format(abs(com[0][self.e_source] - com[1][self.e_source])))
			protruded = ["p_left", "xp_left", "p_right", "xp_right", "average"]
			x2_mod = 5
			if com[0][0] == com[1][0]:
				excess_x = -10 if a["S5"] in ["p_left","xp_left"] else 10
				x1 = {
					"left":self.set_horizontal(com[1][0],self.wide[0]),
					"right":self.set_horizontal(com[1][0],self.wide[1]),
					"midle":self.set_horizontal(com[1][0],int(sum(self.wide)/2)),
					"p_left":self.set_horizontal(com[1][0],self.wide[0]),
					"xp_left": self.set_horizontal(com[1][0], self.wide[0]),
					"p_right": self.set_horizontal(com[1][0], self.wide[1]),
					"xp_right": self.set_horizontal(com[1][0], self.wide[1]),
					"average": self.set_horizontal((com[1][0]+com[0][0])/2, int(sum(self.wide)/2))
				}[a["S5"]]
				x2 = {
					"left":self.set_horizontal(com[1][0],self.wide[0]+x2_mod),
					"right":self.set_horizontal(com[1][0],self.wide[1]-x2_mod),
					"midle":self.set_horizontal(com[1][0],int(sum(self.wide)/2)),
					"p_left":self.set_horizontal(com[1][0],self.wide[0]-2*x2_mod),
					"xp_left": self.set_horizontal(com[1][0], self.wide[0] - 4*x2_mod),
					"p_right": self.set_horizontal(com[1][0], self.wide[1] + 2*x2_mod),
					"xp_right": self.set_horizontal(com[1][0], self.wide[1] + 4*x2_mod),
					"average": self.set_horizontal((com[1][0]+com[0][0])/2, int(sum(self.wide)/2))
				}[a["S5"]]
				label = self.commafy(digit_rounding.format(abs(com[0][self.e_source] - com[1][self.e_source])))
				b = [
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>',  # HORIZONTAL
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" />',  # VERTICAL
					'    <polygon points="{},{} {},{} {},{}" style="fill:{};stroke:{};stroke-width:1"/>',  # TOP Y ARROW
					'    <polygon points="{},{} {},{} {},{}" style="fill:{};stroke:{};stroke-width:1"/>',	# BOTTOM Y ARROW
					'    <text x="{}" y="{}" font-size="16" {} text-anchor="middle" fill="{}">{}</text>',  # Y label
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>'  # HORIZONTAL SHORT

				]
				b[0] = b[0].format(x1, y1, x2 + excess_x, y1, color, width, dashed)
				b[1] = b[1].format(x2, y1 + excess_y, x2, y2 - excess_y, color, width)
				b[2] = b[2].format(x2, y1 + excess_yb, x2 - 4, y1 + excess_y, x2 + 4, y1 + excess_y, color, color)
				b[3] = b[3].format(x2, y2 - excess_yb, x2 - 4, y2 - excess_y, x2 + 4, y2 - excess_y, color, color)
				b[4] = b[4].format(x2 + text_pos[0], int((y2 + y1) / 2),
								   f'transform="rotate({text_pos[1]},{x2 + text_pos[0]},{int((y2 + y1) / 2)})"',
								   color,
								   self.char_care(label))
				b[5] = b[5].format(x1, y2, x2 + excess_x, y2, color, width, dashed)
				protruded = ["p_left", "xp_left", "p_right", "xp_right", "average"]
				self.svg_code.extend(b if a["S5"] in protruded else b[1:5])
			else:
				excess_x = 10 if ordered else -10
				x1 = self.set_horizontal(com[0][0],self.wide[1] if ordered else self.wide[0])
				x2 = {
					"left":self.set_horizontal(com[1][0],self.wide[0]+x2_mod),
					"right":self.set_horizontal(com[1][0],self.wide[1]-x2_mod),
					"midle":self.set_horizontal(com[1][0],int(sum(self.wide)/2)),
					"p_left":self.set_horizontal(com[1][0],self.wide[0]-2*x2_mod),
					"xp_left": self.set_horizontal(com[1][0], self.wide[0] - 4*x2_mod),
					"p_right": self.set_horizontal(com[1][0], self.wide[1] + 2*x2_mod),
					"xp_right": self.set_horizontal(com[1][0], self.wide[1] + 4*x2_mod),
					"average": self.set_horizontal((com[1][0]+com[0][0])/2, int(sum(self.wide)/2))
				}[a["S5"]]

				x3 = {
					"left":None,
					"right":None,
					"midle":None,
					"p_left":self.set_horizontal(com[1][0],self.wide[0]),
					"xp_left": self.set_horizontal(com[1][0], self.wide[0]),
					"p_right": self.set_horizontal(com[1][0], self.wide[1]),
					"xp_right": self.set_horizontal(com[1][0], self.wide[1]),
					"average": self.set_horizontal(com[1][0],self.wide[1] if a["S4"] == "reverse" else self.wide[0])
				}[a["S5"]]

				x4 = {
					"left":None,
					"right":None,
					"midle":None,
					"p_left":x2+excess_x if a["S4"] == "reverse" else x2-excess_x,
					"xp_left": x2+excess_x if a["S4"] == "reverse" else x2-excess_x,
					"p_right":x2-excess_x if a["S4"] == "reverse" else x2+excess_x,
					"xp_right": x2-excess_x if a["S4"] == "reverse" else x2+excess_x,
					"average": x2-excess_x if a["S4"] == "reverse" else x2-excess_x
				}[a["S5"]]

				b = [
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>', # HORIZONTAL
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" />', # VERTICAL
					'    <polygon points="{},{} {},{} {},{}" style="fill:{};stroke:{};stroke-width:1"/>',# TOP Y ARROW
					'    <polygon points="{},{} {},{} {},{}" style="fill:{};stroke:{};stroke-width:1"/>',# BOTTOM Y ARROW
					'    <text x="{}" y="{}" font-size="16" {} text-anchor="middle" fill="{}">{}</text>',#Y label
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>' # HORIZONTAL SHORT

				]
				b[0] = b[0].format(x1,y1,x2+excess_x,y1,color,width,dashed)
				b[1] = b[1].format(x2,y1+excess_y,x2,y2-excess_y,color,width)
				b[2] = b[2].format(x2,y1+excess_yb,x2-4,y1+excess_y,x2+4,y1+excess_y,color,color)
				b[3] = b[3].format(x2,y2-excess_yb,x2-4,y2-excess_y,x2+4,y2-excess_y,color,color)
				b[4] = b[4].format(x2+text_pos[0],int((y2+y1)/2),f'transform="rotate({text_pos[1]},{x2+text_pos[0]},{int((y2+y1)/2)})"',color,self.char_care(label))
				b[5] = b[5].format(x4, y2, x3, y2, color, width, dashed)
				self.svg_code.extend(b if a["S5"] in protruded else b[0:5])

	@functools.lru_cache(maxsize=1)
	def span_dg(self):
		try:
			if not self.span_worthy: return
			if len(self.plot_dict) != 1:
				self.span_worthy = False
				self.msg.append("This software only performs span analysis if one and only one reaction path is ploted\n")
				return
			only_plot = list(self.plot_dict.values())[0]
			r_const = {"kcal/mol": 0.0019872, "kJ/mol": 0.0019872 * 4.184, "eV": 0.0019872 * 0.04336}[self.span["units"]]
			boltz_const = {"kcal/mol":3.29762e-27,"kJ/mol":3.29762e-27 * 4.184,"eV": 3.29762e-27 * 0.04336}[self.span["units"]]
			planck_const = {"kcal/mol":1.58367e-37,"kJ/mol":1.58367e-37 * 4.184, "eV":1.58367e-37 * 0.04336}[self.span["units"]]
			delta_e = only_plot[-1][self.e_source]-only_plot[0][self.e_source]
			kcal_limit = 3.5
			limit = {"kcal/mol": kcal_limit, "kJ/mol": kcal_limit * 4.184,"eV": kcal_limit * 0.04336}[self.span["units"]]
			if not len(only_plot) > 2:
				text = "Need more than two structures for span analysis\n"
				self.msg.append(text)
				self.span_worthy = False; return

			first = min(a[0] for a in only_plot)
			last =  max(a[0] for a in only_plot)
			text = "Analysis assumes structures #{} and #{} have the same geometry,"
			text += " but are energeticaly distiguished by the {} of the reaction.\n"
			self.msg.append(text.format(first, last, "exergonicity" if self.e_source == 3 else "exotermicity"))
			if self.e_source != 3:
				text = "WARNING: Data above should only be used after carefull consideration."
				text += "Enthalpy values were employed in place of Gibbs Free energy\n"
				self.msg.append(text)
			# Is it a TS or INT?
			if self.span["irrespective"] != 1:
				give_up = False
				if not only_plot[0][1] == only_plot[-1][1]:
					text = "#{} and #{} must be the same type: TS or INT\n"
					self.msg.append(text.format(only_plot[0][0],only_plot[-1][0]))
					return
				for i,a in enumerate(only_plot):
					if i == 0 or i+1 == len(only_plot):
						top = only_plot[1][self.e_source] < only_plot[0][self.e_source] and only_plot[-1][self.e_source] > only_plot[-2][self.e_source]
					else:
						top = only_plot[i-1][self.e_source] < a[self.e_source] > only_plot[i+1][self.e_source]
					if top and a[1] == "TS":
						continue
					elif top and a[1] == "INT":
						text = "Are you sure #{} is not a TS? It is directly connected to structures lower in both forward and backwards direction!\n"
						self.msg.append(text.format(a[0]))
					elif not top and a[1] == "TS":
						text = "Are you sure #{} is not an INT? It is directly connected to structure(s) higher in energy!\n"
						self.msg.append(text.format(a[0]))
					elif a[1] not in ["TS","INT"]:
						text = "#{} should be set as either TS or INT, otherwise it will be ploted but excluded from analysis\n"
						self.msg.append(text.format(a[0]))
						give_up = True
				if give_up:
					return
			#TOF
			all_it = []
			for idx_a, a in enumerate(only_plot[:-1]):
				for idx_b,b in enumerate(only_plot[:-1]):
					all_it.append([a, a[self.e_source] - b[self.e_source] - delta_e if idx_b < idx_a else a[self.e_source] - b[self.e_source] , b])

			if self.span["irrespective"] != 1:
				all_it = [a for a in all_it if all([a[0][0] != a[2][0],a[0][1] == "TS", a[2][1] == "INT"])]
			if self.span["irrespective"] == 1:
				text = "WARNING: Data above should only be used after carefull consideration."
				text += "Equations were applied on the assumption that all structures are both intermediates and transition states simultaneously\n"
				self.msg.append(text)
			if all([self.span["irrespective"] != 1, all_it, self.e_source == 3]):
				text = "Ref.: Kozuch, S., Shaik, S. Acc. Chem. Res. 2011, 44, 101.\n"
				self.msg.append(text)
			if all_it:
				denominator = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it)
				tof = (((math.exp(-delta_e / (r_const * self.temperature)) - 1) / denominator) * self.temperature * boltz_const) / planck_const
				all_ts = list(dict.fromkeys([a[0][0] for a in all_it]))
				all_int = list(dict.fromkeys([a[2][0] for a in all_it]))
				x_tof_i = []
				x_tof_ts = []
				for x in all_int:
					a = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it if x == a[2][0]) / denominator
					x_tof_i.append([x, a * 100])
				for x in all_ts:
					a = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it if x == a[0][0]) / denominator
					x_tof_ts.append([x, a * 100])
				self.msg.append("".join("#{:>5}: {:>7.2f}% \n".format(*a) for a in x_tof_i))
				self.msg.append("X(tof) for intermediates:")
				self.msg.append("".join("#{:>5}: {:>7.2f}% \n".format(*a) for a in x_tof_ts))
				self.msg.append("X(tof) for transition states:")
				self.msg.append("TOF as catalytic flux law: {:5e} /h\n".format(tof * 3600))
				if abs(tof) > 1e8:
					self.msg.append("ALERT: Please consider the possibility of diffusion control rates\n")
			#CONDITIONALS FOR SPAN
			if delta_e >= 0:
				self.msg.append("Reaction is {}! Span will not be computed!\n".format("endergonic" if self.e_source == 3 else "endotermic"))
				self.span_worthy = False; return
			if not all(a[1] in ["TS","INT"] for a in only_plot) and self.span["irrespective"] != 1:
				text = "All structures have to be identified as either transition states or intermediates for a strict span analysis."
				text += " Irrestrictive analysis may be caried out by checking the 'irrespective of type(TS/INT)' box."
				text += " No span analysis will be conducted\n"
				self.msg.append(text)
				self.span_worthy = False; return
			# SPAN
			all_it = []
			for idx_a, a in enumerate(only_plot[:-1]):
				for idx_b,b in enumerate(only_plot[:-1]):
					all_it.append([a, a[self.e_source] - b[self.e_source] + delta_e if idx_a <idx_b else a[self.e_source] - b[self.e_source], b])
			if self.span["irrespective"] != 1:
				all_it = [a for a in all_it if all([a[0][0] != a[2][0],a[0][1] == "TS",a[2][1] == "INT"])]
			#for a in all_it: print(a)
			if not all_it:
				self.span_worthy = False
				self.msg.append("No states found!\n")
				return
			span = max(all_it, key=lambda a: a[1])
			if span[1] <= 0:
				self.msg.append("The reaction appears to barrierless and therefore no span will be calculated!\n")
				self.span_worthy = False
				return
			for i,a in enumerate(all_it):
				if 0 <= a[1] > span[1] - limit and any(span[n][0] != a[n][0] for n in [0,2]):
					text = "WARNING: Span from #{} to #{} is only {:.2f} {} lower".format(a[0][0],a[2][0],span[1]-a[1],self.span["units"])
					text += " than #{} to #{} and may influence the rate determining state\n".format(span[0][0],span[2][0])
					self.msg.append(text)
			tof_span = (((math.exp(-span[1]/(r_const*self.temperature))))*self.temperature*boltz_const)/planck_const
			self.msg.append("TOF from span: {:5e} /h\n".format(tof_span*3600))
			return span
		except OverflowError:
			self.msg.append("TOF calculation failed due to floating point overflow! Please review your data and units ({}).".format(self.span["units"]))

	@functools.lru_cache(maxsize=1)
	def graph_span(self):
		if not self.span_worthy: return
		if len(self.plot_dict) != 1: self.span_worthy = False; return
		only_plot = list(self.plot_dict.values())[0]
		tdi_correct = {a: b for a,b in zip(pref.menu_d, [-35,-2,10] if self.plot_np else [-20,-2,10])}
		tdts_correct = {a: b for a, b in zip(pref.menu_d, [-15, 18, 32] if self.plot_np else [-15,2,16])}
		if self.span_dg() is None: self.span_worthy = False; return
		if self.span_worthy:
			data = [[self.span_dg()[0][n] for n in [0,-1]],[self.span_dg()[2][n] for n in [0,-1]]]
			span = self.span_dg()[1]
			data = sorted(data,key=lambda x: x[1])
			delta_e = only_plot[-1][self.e_source]-only_plot[0][self.e_source]
			if self.big_arrow:
				# TDI arrow big
				# print(self.data)
				x = tdi_correct[next(a for a in only_plot if a[0] == data[1][0])[5]] - 40
				p = [self.set_horizontal(data[1][0], 30), data[1][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDI</text>',
					'    <path d=" M {} {} L {} {} L {} {} L {} {} L {} {} L {} {} L {} {} Z "/>']
				arrow_size = [10, -70, -10, -70, -10, -40, -20, -40, 0, -10, 20, -40, 10, -40]
				a[0] = a[0].format(p[0], p[1]+35)
				a[1] = a[1].format(*[int(j/3)+p[0] if i%2 == 0 else int(j/3)+p[1]+70 for i,j in enumerate(arrow_size)])
				self.svg_code.extend(a)
				# TDTS arrow
				x = tdts_correct[next(a for a in only_plot if a[0] == data[0][0])[5]] + 140
				p = [self.set_horizontal(data[0][0], 30), data[0][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDTS</text>',
					'    <path d=" M {} {} L {} {} L {} {} L {} {} L {} {} L {} {} L {} {} Z "/>']
				arrow_size = [10, -10, -10, -10, -10, -40, -20, -40, 0, -70, 20, -40, 10, -40]
				a[0] = a[0].format(p[0], p[1] -25)
				a[1] = a[1].format(*[int(j/3)+p[0] if i%2 == 0 else int(j/3)+p[1]-43 for i,j in enumerate(arrow_size)])
				self.svg_code.extend(a)
			else:
				# TDI arrow
				# print(self.data)
				x = tdi_correct[next(a for a in only_plot if a[0] == data[1][0])[5]] - 40
				p = [self.set_horizontal(data[1][0], 10), data[1][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDI</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">{}</text>']
				a[0] = a[0].format(p[0] + 20, p[1] + 45)
				a[1] = a[1].format(p[0] + 20, p[1] + 63,self.char_care("↓"))
				self.svg_code.extend(a)
				# TDTS arrow
				x = tdts_correct[next(a for a in only_plot if a[0] == data[0][0])[5]] + 140
				p = [self.set_horizontal(data[0][0], 50), data[0][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDTS</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">{}</text>']
				a[0] = a[0].format(p[0] - 20, p[1] -30)
				a[1] = a[1].format(p[0] - 20, p[1] -53,self.char_care("↑"))
				self.svg_code.extend(a)
			# dg and span anotations
			a = [
				'    <text x="120" y="450" text-anchor="left" fill="black">Delta = {}</text>',
				'    <text x="120" y="465" text-anchor="left" fill="black">Span = {}</text>']
			a[0] = a[0].format(self.commafy("{:.2f}".format(delta_e)))
			a[1] = a[1].format(self.commafy("{:.2f}".format(span)))
			self.svg_code.extend(a)

	@functools.lru_cache(maxsize=1)
	def return_svg_code(self):
		if not [a for a in self.data_dict if self.data_dict[a]]:
			self.msg.append("No data!");
			self.graph_frame();
			self.svg_code.append('</svg>');
			return self.svg_code
		self.graph_frame()
		self.graph_grid()
		self.graph_crt_points()
		self.graph_connectors()
		self.graph_comparers()
		if self.span_request: self.span_dg()
		if self.span_request: self.graph_span()
		self.svg_code.append('</svg>')
		return self.svg_code
	def save_svg(self,svg_name):
		svg_name = ".E_profile.svg" if svg_name is None else svg_name
		try:
			with open(os.path.join(os.getcwd(),svg_name), "w") as out_file:
				for line in self.return_svg_code(): out_file.write(line + "\n")
			if svg_name != ".E_profile.svg":
				self.msg.append("Take a look at file {}!".format(svg_name))
			return self.msg
		except FileNotFoundError:
			pass

def initialize():
	set_cw = lambda x: x.grid_columnconfigure(0, weight=1)
	set_rw = lambda x: x.grid_rowconfigure(0, weight=1)
	global pref, note, window, frame2
	pref = Preferences()
	window = tk.Tk()
	#NOTEBOOK
	frame1 = tk.Frame(master=window)
	frame1.grid(column=0,row=0,rowspan=2,sticky="news")
	set_cw(frame1)
	set_rw(frame1)
	note = Note(frame1)
	#GENERAL
	frame2 = tk.Frame(master=window)
	frame2.grid(column=1,row=0,rowspan=1,sticky="news")
	set_cw(frame2)
	set_rw(frame2)
	frame3 = tk.Frame(master=window)
	label = tk.Label(frame3, text="github.com/ricalmang")
	label.grid(column=0,row=0,stick="news")
	frame3.grid(column=1,row=1,rowspan=1)
	menu = GeneralMenu(frame2,name="Actions")
	menu.grid(column=0, row=0,  columnspan=1, pady="0",padx="0", rowspan=1,sticky="news")
	set_cw(window)
	set_rw(window)
	w,h = 925 if sys.platform == "win32" or os.name == "nt" else 1000, 685
	ws = window.winfo_screenwidth() # width of the screen
	hs = window.winfo_screenheight() # height of the screen
	window.minsize(w,h)
	window.maxsize(ws,hs)
	x = int(ws/2 - w/2)
	y = int(hs/2 - h/2)
	window.geometry(f"{w}x{h}+{x}+{y}")
	window.mainloop()
initialize()
